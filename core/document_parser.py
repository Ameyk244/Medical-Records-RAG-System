"""Document parsers — PDF, FHIR JSON, HL7, CSV, DICOM, CCD XML, TXT, DOCX, XLSX."""

from __future__ import annotations

import csv
import io
import json
import re
import statistics
import xml.etree.ElementTree as ET
from pathlib import Path

import pdfplumber
import hl7

from core.chunker import Section


class ParseError(Exception):
    """Raised when an explicit validation check fails inside a parser."""


_FHIR_RESOURCE_TO_SECTION: dict[str, str] = {
    "Patient": "Patient",
    "Condition": "Conditions",
    "MedicationRequest": "Medications",
    "MedicationStatement": "Medications",
    "MedicationAdministration": "Medications",
    "Medication": "Medications",
    "Observation": "Observations",
    "Procedure": "Procedures",
    "AllergyIntolerance": "Allergies",
    "Immunization": "Immunizations",
    "DiagnosticReport": "Diagnostic Reports",
    "Encounter": "Encounters",
    "DocumentReference": "Document References",
    "CarePlan": "Care Plans",
    "Practitioner": "Practitioners",
}

_HL7_SEGMENT_TO_SECTION: dict[str, str] = {
    "MSH": "Message Header",
    "PID": "Patient",
    "PV1": "Visit",
    "AL1": "Allergies",
    "DG1": "Diagnoses",
    "OBX": "Observations",
    "OBR": "Observation Requests",
    "RXA": "Medications",
    "RXE": "Medications",
    "RXO": "Medications",
    "ORC": "Orders",
    "IN1": "Insurance",
    "NK1": "Next of Kin",
    "GT1": "Guarantor",
}


def parse_document(
    source_type: str,
    content: bytes,
    *,
    filename: str | None = None,
) -> list[Section]:
    if source_type == "pdf":
        return _parse_pdf(content)
    elif source_type == "fhir_json":
        return _parse_fhir_json(content)
    elif source_type == "hl7":
        return _parse_hl7(content)
    elif source_type == "csv":
        return _parse_csv(content, filename)
    elif source_type == "dicom_meta":
        return _parse_dicom_meta(content)
    elif source_type == "ccd_xml":
        return _parse_ccd_xml(content)
    elif source_type == "txt":
        return _parse_txt(content)
    elif source_type == "docx":
        return _parse_docx(content)
    elif source_type == "xlsx":
        return _parse_xlsx(content)
    else:
        raise ValueError(f"unsupported source_type: {source_type!r}")


def _parse_pdf(content: bytes) -> list[Section]:
    all_lines: list[tuple[str, float, bool]] = []  # (text, mean_size, is_all_caps)

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(extra_attrs=["size"])
            if not words:
                continue

            # Group words into lines by their `top` coordinate. Words whose
            # `top` values differ by less than 3 points belong to the same line.
            lines_map: dict[float, list[dict]] = {}
            for word in words:
                top = word["top"]
                placed = False
                for key in lines_map:
                    if abs(key - top) < 3:
                        lines_map[key].append(word)
                        placed = True
                        break
                if not placed:
                    lines_map[top] = [word]

            # Emit lines sorted by their representative top coordinate.
            for top_key in sorted(lines_map):
                line_words = lines_map[top_key]
                # Sort left-to-right so joined text reads naturally.
                line_words.sort(key=lambda w: w["x0"])
                text = " ".join(w["text"] for w in line_words)
                sizes = [w["size"] for w in line_words if isinstance(w.get("size"), (int, float))]
                mean_size = statistics.mean(sizes) if sizes else 0.0
                alpha_chars = [c for c in text if c.isalpha()]
                is_all_caps = len(alpha_chars) >= 3 and all(c.isupper() for c in alpha_chars)
                all_lines.append((text, mean_size, is_all_caps))

    if not all_lines:
        raise ParseError("PDF contains no extractable text")

    # Compute body font size as the median of all line sizes; this gives a
    # stable baseline unaffected by a handful of large headings.
    body_size = statistics.median(size for _, size, _ in all_lines if size > 0) if any(
        size > 0 for _, size, _ in all_lines
    ) else 0.0

    def _is_heading(text: str, size: float, is_all_caps: bool) -> bool:
        alpha_chars = [c for c in text if c.isalpha()]
        if len(alpha_chars) < 3:
            return False
        word_count = len(text.split())
        if word_count > 8:
            return False
        return (body_size > 0 and size > body_size * 1.15) or is_all_caps

    sections: list[Section] = []
    current_heading: str | None = None
    current_body: list[str] = []
    found_any_heading = False

    for text, size, is_all_caps in all_lines:
        if _is_heading(text, size, is_all_caps):
            found_any_heading = True
            body_text = " ".join(current_body).strip()
            if body_text:
                sections.append(Section(heading=current_heading, text=body_text))
            current_heading = text
            current_body = []
        else:
            current_body.append(text)

    # Flush the last in-progress section.
    body_text = " ".join(current_body).strip()
    if body_text:
        sections.append(Section(heading=current_heading, text=body_text))

    if not found_any_heading:
        full_text = " ".join(t for t, _, _ in all_lines).strip()
        return [Section(heading=None, text=full_text)]

    return [s for s in sections if s.text]


def _parse_fhir_json(content: bytes) -> list[Section]:
    data = json.loads(content.decode("utf-8"))

    if isinstance(data, dict) and data.get("resourceType") == "Bundle":
        resources = [e["resource"] for e in data.get("entry", []) if "resource" in e]
    elif isinstance(data, list):
        resources = data
    elif isinstance(data, dict) and data.get("resourceType"):
        resources = [data]
    else:
        raise ParseError("FHIR JSON is not a Bundle, resource, or list of resources")

    if not resources:
        raise ParseError("FHIR JSON contains no resources")

    # Preserve insertion order; use dict keyed by section heading.
    section_texts: dict[str, list[str]] = {}

    for resource in resources:
        if not isinstance(resource, dict):
            continue
        resource_type = resource.get("resourceType", "Unknown")
        heading = _FHIR_RESOURCE_TO_SECTION.get(resource_type, resource_type)

        rendered = _render_fhir_resource(resource)
        if not rendered:
            continue

        if heading not in section_texts:
            section_texts[heading] = []
        section_texts[heading].append(rendered)

    return [
        Section(heading=heading, text="\n\n".join(texts))
        for heading, texts in section_texts.items()
    ]


def _render_fhir_resource(resource: dict) -> str:
    # Prefer the FHIR human-readable narrative when present — it is already
    # prose intended for display and requires no structural interpretation.
    narrative_div = resource.get("text", {}).get("div", "")
    if narrative_div:
        stripped = re.sub(r"<[^>]+>", " ", narrative_div)
        collapsed = re.sub(r"\s+", " ", stripped).strip()
        if collapsed:
            return collapsed

    # ------------------------------------------------------------------
    # TODO(Phase 5): This fallback rendering only emits TOP-LEVEL SCALAR
    # fields. Nested CodeableConcept structures — which is where ICD-10
    # codes, SNOMED codes, drug codes, and most clinically actionable
    # data actually live — are NOT unrolled. A FHIR resource with only
    # structured fields and no `text.div` narrative will produce a near-
    # empty section. Phase 5 must add structured-field rendering before
    # this parser is used on production FHIR exports.
    # ------------------------------------------------------------------
    _skip = {"resourceType", "id", "meta", "text"}
    parts: list[str] = []
    for key, value in resource.items():
        if key in _skip or key.startswith("_"):
            continue
        if isinstance(value, (str, int, float, bool)):
            parts.append(f"{key}: {value}")
    return ". ".join(parts)


def _parse_hl7(content: bytes) -> list[Section]:
    # HL7 v2 has historical 8-bit encoding conventions; latin-1 is a safe
    # superset of ASCII that never raises a decode error on raw byte streams.
    text = content.decode("latin-1")

    # Normalize all line-ending variants to the canonical HL7 segment
    # terminator \r so the `hl7` library can parse them reliably.
    normalized = text.replace("\r\n", "\r").replace("\n", "\r")

    try:
        message = hl7.parse(normalized)
    except Exception as exc:
        raise ParseError(f"failed to parse HL7 message: {exc}") from exc

    # Collect all segments from the parsed message.
    segments = list(message)
    if not segments:
        raise ParseError("HL7 message contains no segments")

    section_texts: dict[str, list[str]] = {}
    for segment in segments:
        segment_code = str(segment[0])
        heading = _HL7_SEGMENT_TO_SECTION.get(segment_code, segment_code)
        if heading not in section_texts:
            section_texts[heading] = []
        section_texts[heading].append(str(segment))

    return [
        Section(heading=heading, text="\n".join(lines))
        for heading, lines in section_texts.items()
    ]


def _parse_csv(content: bytes, filename: str | None) -> list[Section]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        raise ParseError("CSV contains no rows")

    # Derive a human-readable heading from the filename when provided, so
    # the section label reflects what the file actually contains.
    if filename:
        stem = Path(filename).stem
        heading = stem.replace("_", " ").replace("-", " ").title()
    elif reader.fieldnames:
        heading = reader.fieldnames[0]
    else:
        heading = "Tabular Data"

    row_lines: list[str] = []
    for row in rows:
        parts = [f"{col}: {val}" for col, val in row.items() if val not in (None, "")]
        if parts:
            row_lines.append(", ".join(parts))

    body = "\n".join(row_lines)
    return [Section(heading=heading, text=body)]


def _parse_dicom_meta(content: bytes) -> list[Section]:
    # Lazy import — pydicom is large and only needed for DICOM files.
    try:
        import pydicom
    except ImportError as e:
        raise ParseError("pydicom is not installed") from e

    try:
        ds = pydicom.dcmread(io.BytesIO(content), stop_before_pixels=True, force=True)
    except Exception as e:
        raise ParseError(f"failed to parse DICOM: {e}") from e

    _DICOM_TAGS = [
        "PatientID", "PatientName", "PatientBirthDate", "PatientSex",
        "StudyDate", "StudyTime", "StudyDescription", "AccessionNumber",
        "Modality", "BodyPartExamined", "SeriesDescription", "SeriesNumber",
        "Manufacturer", "ManufacturerModelName", "InstitutionName",
    ]

    lines: list[str] = []
    for tag in _DICOM_TAGS:
        value = getattr(ds, tag, None)
        if value is not None:
            lines.append(f"{tag}: {value}")

    if not lines:
        raise ParseError("DICOM file contains no recognisable metadata")

    return [Section(heading="DICOM Metadata", text="\n".join(lines))]


def _parse_ccd_xml(content: bytes) -> list[Section]:
    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        raise ParseError(f"failed to parse CCD XML: {e}") from e

    # Strip namespace prefix so we can match tag names regardless of which
    # HL7 namespace prefix the document author chose.
    def _local(tag: str) -> str:
        return tag.split("}", 1)[-1]

    sections: list[Section] = []
    for node in root.iter():
        if _local(node.tag) == "section":
            title: str | None = None
            body_parts: list[str] = []
            for child in node:
                local = _local(child.tag)
                if local == "title":
                    title = (child.text or "").strip() or None
                elif local == "text":
                    # Recursively flatten all nested text nodes — CCD <text>
                    # can contain tables, lists, and other inline markup.
                    raw = "".join(child.itertext())
                    collapsed = re.sub(r"\s+", " ", raw).strip()
                    if collapsed:
                        body_parts.append(collapsed)
            body = " ".join(body_parts).strip()
            if body:
                sections.append(Section(heading=title, text=body))

    if sections:
        return sections

    # Fallback: no <section> elements found — return all text from the document.
    full_text = re.sub(r"\s+", " ", "".join(root.itertext())).strip()
    if not full_text:
        raise ParseError("CCD XML contains no readable text")
    return [Section(heading=None, text=full_text)]


def _parse_txt(content: bytes) -> list[Section]:
    # utf-8-sig strips the BOM when present; latin-1 handles legacy encodings
    # without ever raising a decode error on arbitrary byte values.
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    if not text.strip():
        raise ParseError("text file is empty")

    paragraphs = re.split(r"\n\s*\n", text)

    def _is_heading_line(line: str) -> bool:
        alpha = [c for c in line if c.isalpha()]
        if len(alpha) < 3:
            return False
        if len(line.split()) > 8:
            return False
        return all(c.isupper() for c in alpha)

    sections: list[Section] = []
    current_heading: str | None = None
    current_body: list[str] = []
    found_any_heading = False

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        lines = para.splitlines()
        first_line = lines[0].strip()
        if _is_heading_line(first_line):
            found_any_heading = True
            body_text = " ".join(current_body).strip()
            if body_text:
                sections.append(Section(heading=current_heading, text=body_text))
            current_heading = first_line
            rest = " ".join(l.strip() for l in lines[1:]).strip()
            current_body = [rest] if rest else []
        else:
            current_body.append(para)

    body_text = " ".join(current_body).strip()
    if body_text:
        sections.append(Section(heading=current_heading, text=body_text))

    if not found_any_heading:
        collapsed = re.sub(r"\s+", " ", text).strip()
        return [Section(heading=None, text=collapsed)]

    return [s for s in sections if s.text]


def _parse_docx(content: bytes) -> list[Section]:
    # Lazy import — python-docx is large and only needed for DOCX files.
    try:
        from docx import Document as DocxDocument
    except ImportError as e:
        raise ParseError("python-docx is not installed") from e

    try:
        doc = DocxDocument(io.BytesIO(content))
    except Exception as e:
        raise ParseError(f"failed to parse DOCX: {e}") from e

    paragraphs = doc.paragraphs
    if not paragraphs:
        raise ParseError("DOCX contains no readable paragraphs")

    _HEADING_STYLES = {"Heading 1", "Heading 2", "Heading 3"}

    sections: list[Section] = []
    current_heading: str | None = None
    current_body: list[str] = []
    found_any_heading = False

    for p in paragraphs:
        text = p.text.strip()
        if not text:
            continue
        if p.style.name in _HEADING_STYLES:
            found_any_heading = True
            body_text = " ".join(current_body).strip()
            if body_text:
                sections.append(Section(heading=current_heading, text=body_text))
            current_heading = text
            current_body = []
        else:
            current_body.append(text)

    body_text = " ".join(current_body).strip()
    if body_text:
        sections.append(Section(heading=current_heading, text=body_text))

    if not found_any_heading:
        all_text = " ".join(p.text.strip() for p in paragraphs if p.text.strip())
        return [Section(heading=None, text=all_text)]

    return [s for s in sections if s.text]


def _parse_xlsx(content: bytes) -> list[Section]:
    # Lazy import — openpyxl is large and only needed for XLSX files.
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ParseError("openpyxl is not installed") from e

    try:
        # data_only=True returns evaluated cell values instead of formula
        # strings, which is what we want for clinical data extraction.
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"failed to parse XLSX: {e}") from e

    sections: list[Section] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        row_lines: list[str] = []
        for row in rows[1:]:
            parts = [
                f"{headers[i]}: {val}"
                for i, val in enumerate(row)
                if i < len(headers) and val not in (None, "")
            ]
            if parts:
                row_lines.append(", ".join(parts))

        if row_lines:
            sections.append(Section(heading=sheet.title, text="\n".join(row_lines)))

    if not sections:
        raise ParseError("XLSX contains no rows")

    return sections
